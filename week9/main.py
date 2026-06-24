import anthropic
import json
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import HTMLResponse, StreamingResponse
from pydantic import BaseModel
from datetime import datetime

app = FastAPI()
client = anthropic.Anthropic()


# ─── Section 2: VALUATION DATE LOGIC ────────────────────────────────────────────

def parse_valuation_date(date_str: str) -> dict:
    val_date = datetime.strptime(date_str, "%Y-%m-%d")
    year_end = datetime(val_date.year, 12, 31)
    stub_months = (year_end.year - val_date.year) * 12 + (year_end.month - val_date.month)

    if stub_months == 0:
        stub_discount_period = 0
        first_forecast_year = val_date.year + 1
    else:
        stub_discount_period = round((stub_months / 2) / 12, 4)
        first_forecast_year = val_date.year + 1

    return {
        "val_date": val_date,
        "val_year": val_date.year,
        "stub_months": stub_months,
        "stub_discount_period": stub_discount_period,
        "first_forecast_year": first_forecast_year,
        "stub_label": f"YTG {val_date.year}"
    }

def get_discount_period(n: int, stub_months: int) -> float:
    return round((stub_months / 12) + (n - 0.5), 4)

def get_tv_discount_period(n_years: int, stub_months: int) -> float:
    return round((stub_months / 12) + n_years, 4)



# ─── Section 3: SMART COLUMN CLASSIFIER ────────────────────────────────────────

def classify_columns(wb: openpyxl.Workbook, val_date_info: dict) -> dict:
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_row_idx = None
        headers = []

        for i, row in enumerate(rows[:20]):
            non_empty = [str(c) for c in row if c is not None]
            time_keywords = ['LTM', 'YTD', 'YTG', 'FY', 'CY', 'TTM', 'H1', 'H2',
                           str(val_date_info['val_year']),
                           str(val_date_info['first_forecast_year'])]
            if len(non_empty) >= 2 and any(
                any(kw in str(c).upper() for kw in time_keywords)
                for c in non_empty
            ):
                header_row_idx = i
                headers = list(row)
                break

        if header_row_idx is None:
            continue

        headers_str = [str(h) if h else "" for h in headers]
        val_year = val_date_info['val_year']

        response = client.messages.create(
            model="claude-opus-4-5",
            max_tokens=800,
            messages=[{"role": "user", "content": f"""
Classify these column headers from a financial model.
Valuation date: {val_date_info['val_date'].strftime('%d %B %Y')}
Current year: {val_year}

Headers (with their index): {list(enumerate(headers_str))}

Classify each non-empty header as one of:
- "historical": actual data from years before {val_year}
- "ytd": year-to-date actuals in {val_year}
- "ytg": year-to-go forecast for remaining months of {val_year}
- "ltm": last twelve months
- "forecast": forecast years after {val_year}
- "tv": terminal value column
- "label": the first column containing row labels
- "other": anything else

Return ONLY valid JSON:
{{"sheet": "{sheet_name}", "columns": [{{"index": 0, "header": "...", "type": "...", "year": null}}]}}
"""}]
        )

        text = response.content[0].text.strip()
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
            text = text.strip()

        try:
            result[sheet_name] = json.loads(text)
            result[sheet_name]["header_row_idx"] = header_row_idx
            result[sheet_name]["headers"] = headers_str
        except:
            result[sheet_name] = {"error": "Could not parse columns"}

    return result


# ─── Section 4: WACC TAB CREATOR ────────────────────────────────────────────

def add_wacc_tab(wb: openpyxl.Workbook) -> None:
    if "WACC" in wb.sheetnames:
        del wb["WACC"]

    ws = wb.create_sheet("WACC", 0)

    # Styles
    header_fill = PatternFill("solid", fgColor="1F3864")
    input_fill  = PatternFill("solid", fgColor="DCE6F1")
    calc_fill   = PatternFill("solid", fgColor="E2EFDA")
    wacc_fill   = PatternFill("solid", fgColor="86BC25")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    label_font  = Font(size=10)
    bold_font   = Font(bold=True, size=10)
    wacc_font   = Font(bold=True, size=12, color="FFFFFF")

    def write_row(row, label, formula=None, is_input=False, unit="", bold=False):
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = bold_font if bold else label_font
        ws[f"D{row}"] = unit
        if formula:
            ws[f"C{row}"] = formula
            ws[f"C{row}"].fill = calc_fill
            ws[f"C{row}"].font = bold_font
        elif is_input:
            ws[f"C{row}"].fill = input_fill

    # Title
    ws["A1"] = "WACC — CAPM Analysis"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:D1")

    # Section: Cost of Equity
    ws["A3"] = "COST OF EQUITY"
    ws["A3"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill("solid", fgColor="2E5DA8")
    ws.merge_cells("A3:D3")

    write_row(4,  "Risk-Free Rate",                  is_input=True,  unit="%")
    write_row(5,  "Equity Risk Premium",              is_input=True,  unit="%")
    write_row(6,  "Beta (Relevered)",                 is_input=True,  unit="x")
    write_row(7,  "Size Premium",                     is_input=True,  unit="%")
    write_row(8,  "Company-Specific Risk Premium",    is_input=True,  unit="%")
    write_row(9,  "Cost of Equity",  "=C4+C6*C5+C7+C8",             unit="%", bold=True)

    # Section: Cost of Debt
    ws["A11"] = "COST OF DEBT"
    ws["A11"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A11"].fill = PatternFill("solid", fgColor="2E5DA8")
    ws.merge_cells("A11:D11")

    write_row(12, "Pre-Tax Cost of Debt",             is_input=True,  unit="%")
    write_row(13, "Tax Rate",                         is_input=True,  unit="%")
    write_row(14, "After-Tax Cost of Debt", "=C12*(1-C13/100)",          unit="%", bold=True)

    # Section: Capital Structure
    ws["A16"] = "CAPITAL STRUCTURE"
    ws["A16"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A16"].fill = PatternFill("solid", fgColor="2E5DA8")
    ws.merge_cells("A16:D16")

    write_row(17, "Equity Weight",                    is_input=True,  unit="%")
    write_row(18, "Debt Weight",          "=100-C17",                   unit="%", bold=True)

    # WACC Result
    ws["A20"] = "WACC"
    ws["A20"].font = wacc_font
    ws["A20"].fill = wacc_fill
    ws["C20"] = "=C17/100*C9+(100-C17)/100*C14"
    ws["C20"].font = wacc_font
    ws["C20"].fill = wacc_fill
    ws["D20"] = "%"
    ws["D20"].font = wacc_font
    ws["D20"].fill = wacc_fill
    ws.merge_cells("A20:B20")

    # Note
    ws["A22"] = "Note: Blue cells = user inputs. Green cells = calculated. Fill blue cells to update WACC."
    ws["A22"].font = Font(italic=True, size=9, color="666666")

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 5


# ─── Section 5: HTML FRONT END ────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
def home():
    return """
<!DOCTYPE html>
<html>
<head>
    <title>Deloitte Valuation Tool</title>
    <style>
        body { font-family: Arial, sans-serif; max-width: 750px; margin: 40px auto; padding: 0 20px; background: #f5f5f5; }
        h1 { color: #86BC25; }
        .card { background: white; border-radius: 8px; padding: 24px; margin-bottom: 20px; border: 1px solid #ddd; }
        h3 { color: #1F3864; margin-top: 0; }
        label { display: block; font-weight: bold; margin: 12px 0 4px 0; font-size: 13px; color: #333; }
        input, select { width: 100%; padding: 8px; border: 1px solid #ccc; border-radius: 4px; box-sizing: border-box; font-size: 13px; }
        .upload-area { border: 2px dashed #86BC25; border-radius: 8px; padding: 30px; text-align: center; background: #f9fef0; }
        button { background: #86BC25; color: white; padding: 12px 30px; border: none; border-radius: 4px; font-size: 15px; cursor: pointer; width: 100%; margin-top: 16px; }
        #status { display: none; background: #e8f5e9; border-left: 4px solid #86BC25; padding: 12px 16px; border-radius: 4px; margin-top: 16px; }
        #error  { display: none; background: #fdecea; border-left: 4px solid #c62828; padding: 12px 16px; border-radius: 4px; margin-top: 16px; }
    </style>
</head>
<body>
    <h1>Deloitte Valuation Tool</h1>

    <div class="card">
        <h3>Engagement Details</h3>

        <label>Company Name</label>
        <input type="text" id="company_name" placeholder="e.g. GEMS Education">

        <label>Sector</label>
        <select id="sector">
            <option value="education">Education</option>
            <option value="real estate">Real Estate</option>
            <option value="healthcare">Healthcare</option>
            <option value="technology">Technology</option>
            <option value="retail">Retail</option>
            <option value="hospitality">Hospitality</option>
            <option value="construction">Construction</option>
            <option value="other">Other</option>
        </select>

        <label>Valuation Date</label>
        <input type="date" id="valuation_date" value="2025-06-30">
    </div>

    <div class="card">
        <h3>Client Financial Data</h3>
        <div class="upload-area">
            <p>Upload client Excel file (.xlsx)</p>
            <input type="file" id="fileInput" accept=".xlsx">
        </div>
    </div>

    <button onclick="uploadFile()">Build Valuation Model</button>

    <div id="status"></div>
    <div id="error"></div>

    <script>
        async function uploadFile() {
            const file = document.getElementById('fileInput').files[0];
            const company = document.getElementById('company_name').value;
            const sector = document.getElementById('sector').value;
            const valDate = document.getElementById('valuation_date').value;

            if (!file) { alert('Please select an Excel file'); return; }
            if (!company) { alert('Please enter company name'); return; }
            if (!valDate) { alert('Please enter valuation date'); return; }

            document.getElementById('status').style.display = 'block';
            document.getElementById('status').innerHTML = '⏳ Reading client data and building valuation model...';
            document.getElementById('error').style.display = 'none';

            const formData = new FormData();
            formData.append('file', file);
            formData.append('company_name', company);
            formData.append('sector', sector);
            formData.append('valuation_date', valDate);

            try {
                const response = await fetch('/upload', {
                    method: 'POST',
                    body: formData
                });

                if (!response.ok) {
                    const err = await response.text();
                    throw new Error(err);
                }

                const blob = await response.blob();
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = company.replace(/\s+/g, '_') + '_Valuation_Model.xlsx';
                a.click();

                document.getElementById('status').innerHTML =
                    '✅ Valuation model ready — check your downloads folder. Open it in Excel and fill in the blue WACC cells.';
            } catch (err) {
                document.getElementById('status').style.display = 'none';
                document.getElementById('error').style.display = 'block';
                document.getElementById('error').innerHTML = '❌ Error: ' + err.message;
            }
        }
    </script>
</body>
</html>
"""


# ─── Section 6: FIND FINANCIAL ROWS ────────────────────────────────────────────

# ─── Section 7: FIND FINANCIAL ROWS ────────────────────────────────────────────
def find_financial_rows(wb: openpyxl.Workbook, column_map: dict) -> dict:
    """
    Scan each source sheet and return row numbers for key financial line items.
    Checks the first 3 columns of each row to find the label (handles sheets
    where column A is empty and labels sit in column B or C).
    Returns: {sheet_name: {metric: {row, label}}}
    """
    search_terms = {
        "revenue":    ["revenue", "net revenue", "total revenue", "sales", "turnover"],
        "ebitda":     ["ebitda"],
        "ebit":       ["ebit", "operating profit", "operating income"],
        "da":         ["depreciation", "d&a", "dep & amort",
                       "depreciation and amortization", "add: depreciation",
                       "add back depreciation"],
        "capex":      ["capex", "capital expenditure", "capital expenditures",
                       "purchase of ppe", "acquisition of property",
                       "less: capex", "investment in ppe"],
        "wc_balance": ["nwc balance", "net working capital balance", 
                       "closing nwc"],
    }
    exclude = ["margin", "growth", "% of", "per share", "ratio", "change in",
               "movement", "increase", "decrease"]

    financial_map = {}

    for sheet_name, info in column_map.items():
        ws = wb[sheet_name]
        header_row_idx = info.get("header_row_idx", 0)   # 0-based index
        found = {}
        labels_seen = []

        # Start scanning from the row AFTER the header row (convert to 1-based)
        for row in ws.iter_rows(min_row=header_row_idx + 2):
            # Check first 3 columns to find the text label
            label = None
            for col_idx in range(min(3, len(row))):
                cell = row[col_idx]
                if cell.value and isinstance(cell.value, str) and cell.value.strip():
                    label = cell.value.strip()
                    break

            if not label:
                continue

            label_lower = label.lower()
            labels_seen.append(label_lower)

            if any(ex in label_lower for ex in exclude):
                continue

            row_number = row[0].row    # actual Excel row number for formula references

            for key, keywords in search_terms.items():
                if key not in found:
                    if any(kw in label_lower for kw in keywords):
                        if key == "ebit" and "ebitda" in label_lower:
                            continue
                        found[key] = {"row": row_number, "label": label}

        print(f"DEBUG '{sheet_name}' labels: {labels_seen[:20]}")

        if found:
            financial_map[sheet_name] = found

    return financial_map


# ─── Section 7: ADD DCF TAB ─────────────────────────────────────────────────────
def add_dcf_tab(wb: openpyxl.Workbook, val_date_info: dict,
                column_map: dict, financial_map: dict, company_name: str):

    # ── Colour palette ──
    DARK_BLUE    = "1F3864"
    MED_BLUE     = "2E5DA8"
    INPUT_BLUE   = "DCE6F1"
    CALC_GREEN   = "E2EFDA"
    RESULT_GREEN = "86BC25"
    WHITE        = "FFFFFF"

    # ── Pick primary source sheet (first with forecast columns) ──
    src_sheet   = None
    src_columns = []
    for sheet_name, info in column_map.items():
        cols = info.get("columns", [])
        if any(c.get("type") in ("forecast", "ytg", "stub") for c in cols):
            src_sheet   = sheet_name
            src_columns = cols
            break
    
    # Compute Excel column letter from index for each column   
    for col in src_columns:                                    
        col["col_letter"] = get_column_letter(col["index"] + 1)   
    
    if not src_sheet:
        return

    row_map = financial_map.get(src_sheet, {})

    # ── Sort ALL source columns by their letter (A→Z) ──
    
    all_sorted = sorted(src_columns,
                        key=lambda c: column_index_from_string(c.get("col_letter", "A")))

    # ── Build periods list (stub + forecast only) ──
    stub_months = val_date_info["stub_months"]
    period_cols = [c for c in all_sorted
                   if c.get("type") in ("ytg", "stub", "forecast")]

    periods = []
    forecast_count = 0
    for i, col in enumerate(period_cols):
        dcf_col = get_column_letter(i + 2)   # B, C, D, …

        # Discount period
        if col.get("type") in ("ytg", "stub"):
            disc  = val_date_info["stub_discount_period"]
            label = val_date_info["stub_label"]
        else:
            forecast_count += 1
            disc  = get_discount_period(forecast_count, stub_months)
            label = f"FY {col['year']}"

        # Previous source column (for ΔWC = WC_curr − WC_prev)
        col_idx_in_all = next((j for j, c in enumerate(all_sorted)
                               if c.get("col_letter") == col.get("col_letter")), None)
        prev_src_col = (all_sorted[col_idx_in_all - 1]["col_letter"]
                        if col_idx_in_all and col_idx_in_all > 0 else "")

        periods.append({
            "label":        label,
            "dcf_col":      dcf_col,
            "src_col":      col.get("col_letter", ""),
            "prev_src_col": prev_src_col,
            "discount":     disc,
            "type":         col.get("type"),
            "year":         col.get("year"),
        })

    if not periods:
        return

    n_forecast   = sum(1 for p in periods if p["type"] not in ("ytg", "stub"))
    tv_disc      = get_tv_discount_period(n_forecast, stub_months)
    last_col     = next((p["dcf_col"] for p in reversed(periods)
                         if p["type"] not in ("ytg","stub")), periods[-1]["dcf_col"])

    # ── Helper: cross-sheet formula reference ──
    def src_ref(metric, period):
        if metric not in row_map:
            return ""
        rn = row_map[metric]["row"]
        sc = period["src_col"]
        return f"='{src_sheet}'!{sc}{rn}" if sc else ""

    # ── Helper: ΔWC formula (absolute WC curr − WC prev) ──
    def wc_delta_ref(period):
        if "wc_balance" not in row_map:
            return ""
        rn       = row_map["wc_balance"]["row"]
        curr_col = period["src_col"]
        prev_col = period["prev_src_col"]
        if not curr_col or not prev_col:
            return ""
        return f"='{src_sheet}'!{curr_col}{rn}-'{src_sheet}'!{prev_col}{rn}"
    
    def abs_src_ref(metric, period):
        """Always positive — handles both positive and negative source conventions."""
        ref = src_ref(metric, period)
        if ref and ref.startswith("="):
            return "=ABS(" + ref[1:] + ")"
        return ref

    # ── Create / replace tab ──
    if "DCF Valuation" in wb.sheetnames:
        del wb["DCF Valuation"]
    ws = wb.create_sheet("DCF Valuation", 1)

    # ── Column widths ──
    ws.column_dimensions["A"].width = 42
    for p in periods:
        ws.column_dimensions[p["dcf_col"]].width = 16

    # ── Row-writing helpers ──
    def fill_cell(row, col_letter, value, bold=False, fg="000000", bg=WHITE,
                  align="center", fmt='#,##0.0'):
        c = ws.cell(row=row,
                    column=column_index_from_string(col_letter),
                    value=value)
        c.font      = Font(name="Calibri", bold=bold, color=fg, size=10)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center",
                                indent=(1 if align == "left" else 0))
        if isinstance(value, str) and value.startswith("="):
            c.number_format = fmt
        elif isinstance(value, (int, float)):
            c.number_format = fmt
        return c

    def section_header(row, text, color):
        ws.row_dimensions[row].height = 18
        c = fill_cell(row, "A", text, bold=True, fg=WHITE, bg=color, align="left")
        if periods:
            ws.merge_cells(f"A{row}:{periods[-1]['dcf_col']}{row}")

    def data_row(row, label, vals_by_dcf_col, style="calc"):
        bg_map = {"input": INPUT_BLUE, "calc": CALC_GREEN,
                  "result": RESULT_GREEN, "plain": WHITE}
        bg   = bg_map.get(style, WHITE)
        bold = style == "result"
        fg   = WHITE if style == "result" else "000000"
        ws.row_dimensions[row].height = 16
        fill_cell(row, "A", label, bold=bold, fg=fg, bg=bg, align="left")
        for p in periods:
            fill_cell(row, p["dcf_col"],
                      vals_by_dcf_col.get(p["dcf_col"], ""),
                      bold=bold, fg=fg, bg=bg)

    # ─────────────────────────────────────────────────────────────────────────────
    r = 1

    # Title
    ws.row_dimensions[r].height = 24
    fill_cell(r, "A", f"DCF VALUATION  —  {company_name.upper()}",
              bold=True, fg=WHITE, bg=DARK_BLUE, align="left", fmt="@")
    ws.merge_cells(f"A{r}:{periods[-1]['dcf_col']}{r}")
    r += 1

    # Subtitle
    ws.row_dimensions[r].height = 15
    sub = (f"Valuation Date: {val_date_info['val_date'].strftime('%d %B %Y')}"
           f"   |   Source: {src_sheet}   |   Currency: USD'm")
    fill_cell(r, "A", sub, fg=WHITE, bg=MED_BLUE, align="left", fmt="@")
    ws.merge_cells(f"A{r}:{periods[-1]['dcf_col']}{r}")
    r += 1

    r += 1  # blank

    # Period headers
    ws.row_dimensions[r].height = 20
    fill_cell(r, "A", "", bg=MED_BLUE)
    for p in periods:
        fill_cell(r, p["dcf_col"], p["label"], bold=True, fg=WHITE, bg=MED_BLUE)
    HEADER_ROW = r;  r += 1

    # Discount periods
    data_row(r, "Discount Period",
             {p["dcf_col"]: p["discount"] for p in periods}, style="plain")
    DISC_ROW = r;  r += 1

    r += 1  # blank

    # ── SECTION A: INCOME SUMMARY ──────────────────────────────────────────────
    section_header(r, "INCOME STATEMENT SUMMARY", MED_BLUE);  r += 1

    data_row(r, "Revenue (USD'm)",
             {p["dcf_col"]: src_ref("revenue", p) for p in periods})
    r += 1

    data_row(r, "EBITDA (USD'm)",
             {p["dcf_col"]: src_ref("ebitda", p) for p in periods})
    EBITDA_ROW = r;  r += 1

    data_row(r, "EBIT (USD'm)",
             {p["dcf_col"]: src_ref("ebit", p) for p in periods})
    EBIT_ROW = r;  r += 1

    r += 1  # blank

    # ── SECTION B: FCFF DERIVATION ─────────────────────────────────────────────
    section_header(r, "FREE CASH FLOW TO FIRM  (FCFF)  DERIVATION", MED_BLUE);  r += 1

    data_row(r, "EBIT (USD'm)",
             {p["dcf_col"]: f"={p['dcf_col']}{EBIT_ROW}" for p in periods})
    EBIT2_ROW = r;  r += 1

    data_row(r, "Tax Rate (%)  — enter as whole number  (e.g. 9 for 9%)",
             {p["dcf_col"]: 9 for p in periods}, style="input")
    TAX_ROW = r;  r += 1

    data_row(r, "NOPAT  =  EBIT × (1 − Tax Rate)",
             {p["dcf_col"]: f"={p['dcf_col']}{EBIT2_ROW}*(1-{p['dcf_col']}{TAX_ROW}/100)"
              for p in periods})
    NOPAT_ROW = r;  r += 1

    data_row(r, "Add: Depreciation & Amortisation (USD'm)",
             {p["dcf_col"]: abs_src_ref("da", p) for p in periods})
    DA_ROW = r;  r += 1

    data_row(r, "Less: Capital Expenditure (USD'm)",
             {p["dcf_col"]: abs_src_ref("capex", p) for p in periods})
    CAPEX_ROW = r;  r += 1

    data_row(r, "Less: Change in Working Capital  [WC(curr) − WC(prior)]  (USD'm)",
             {p["dcf_col"]: wc_delta_ref(p) for p in periods})
    WC_ROW = r;  r += 1

    data_row(r, "FCFF  =  NOPAT + D&A − CapEx − ΔWC  (USD'm)",
             {p["dcf_col"]:
              f"={p['dcf_col']}{NOPAT_ROW}+{p['dcf_col']}{DA_ROW}"
              f"-{p['dcf_col']}{CAPEX_ROW}-{p['dcf_col']}{WC_ROW}"
              for p in periods},
             style="result")
    FCFF_ROW = r;  r += 1

    r += 1  # blank

    # ── SECTION C: DCF ANALYSIS ────────────────────────────────────────────────
    section_header(r, "DCF ANALYSIS", MED_BLUE);  r += 1

    data_row(r, "Discount Period",
             {p["dcf_col"]: p["discount"] for p in periods}, style="plain")
    DISC2_ROW = r;  r += 1

    data_row(r, "Discount Factor  [ 1 / (1 + WACC)ⁿ ]",
             {p["dcf_col"]:
              f"=1/(1+'WACC'!C20/100)^{p['dcf_col']}{DISC2_ROW}"
              for p in periods})
    DF_ROW = r;  r += 1

    data_row(r, "PV of FCFF (USD'm)",
             {p["dcf_col"]:
              f"={p['dcf_col']}{FCFF_ROW}*{p['dcf_col']}{DF_ROW}"
              for p in periods})
    PV_ROW = r;  r += 1

    r += 1  # blank

    # Sum of PVs — single value cell in column B
    ws.row_dimensions[r].height = 16
    fill_cell(r, "A", "Sum of PV of FCFFs (USD'm)",
              bold=True, bg=CALC_GREEN, align="left")
    sum_start = periods[0]["dcf_col"]
    sum_end   = periods[-1]["dcf_col"]
    fill_cell(r, "B", f"=SUM({sum_start}{PV_ROW}:{sum_end}{PV_ROW})",
              bold=True, bg=CALC_GREEN)
    SUM_PV_ROW = r;  r += 1

    r += 1  # blank

    # ── SECTION D: TERMINAL VALUE ──────────────────────────────────────────────
    section_header(r, "TERMINAL VALUE", MED_BLUE);  r += 1

    # TV discount period
    ws.row_dimensions[r].height = 16
    fill_cell(r, "A", "TV Discount Period", bg=WHITE, align="left")
    fill_cell(r, "B", tv_disc, bg=WHITE)
    TV_DISC_ROW = r;  r += 1

    # Terminal growth rate input
    ws.row_dimensions[r].height = 16
    fill_cell(r, "A",
              "Terminal Growth Rate (%)  — enter as whole number  (e.g. 3 for 3%)",
              bg=INPUT_BLUE, align="left")
    fill_cell(r, "B", 3, bg=INPUT_BLUE)
    TGR_ROW = r;  r += 1

    # GGM TV
    ws.row_dimensions[r].height = 16
    fill_cell(r, "A",
              "Gordon Growth Model TV  =  FCFFₙ × (1+g) / (WACC − g)",
              bg=CALC_GREEN, align="left")
    fill_cell(r, "B",
              f"={last_col}{FCFF_ROW}*(1+B{TGR_ROW}/100)"
              f"/('WACC'!C20/100-B{TGR_ROW}/100)",
              bg=CALC_GREEN)
    GGM_TV_ROW = r;  r += 1

    # PV of GGM TV
    ws.row_dimensions[r].height = 16
    fill_cell(r, "A", "PV of GGM Terminal Value (USD'm)",
              bold=True, bg=CALC_GREEN, align="left")
    fill_cell(r, "B",
              f"=B{GGM_TV_ROW}/(1+'WACC'!C20/100)^B{TV_DISC_ROW}",
              bold=True, bg=CALC_GREEN)
    PV_GGM_ROW = r;  r += 1

    r += 1  # blank

    # Exit multiple input
    ws.row_dimensions[r].height = 16
    fill_cell(r, "A",
              "Exit EV/EBITDA Multiple (×)  — enter as number  (e.g. 8 for 8×)",
              bg=INPUT_BLUE, align="left")
    fill_cell(r, "B", 8, bg=INPUT_BLUE)
    MULT_ROW = r;  r += 1

    # Exit Multiple TV
    ws.row_dimensions[r].height = 16
    fill_cell(r, "A",
              f"Exit Multiple TV  =  EBITDAₙ ({last_col}) × Multiple",
              bg=CALC_GREEN, align="left")
    fill_cell(r, "B",
              f"={last_col}{EBITDA_ROW}*B{MULT_ROW}",
              bg=CALC_GREEN)
    EXIT_TV_ROW = r;  r += 1

    # PV of Exit TV
    ws.row_dimensions[r].height = 16
    fill_cell(r, "A", "PV of Exit Multiple Terminal Value (USD'm)",
              bold=True, bg=CALC_GREEN, align="left")
    fill_cell(r, "B",
              f"=B{EXIT_TV_ROW}/(1+'WACC'!C20/100)^B{TV_DISC_ROW}",
              bold=True, bg=CALC_GREEN)
    PV_EXIT_ROW = r;  r += 1

    r += 1  # blank

    # ── SECTION E: ENTERPRISE VALUE ────────────────────────────────────────────
    section_header(r, "ENTERPRISE VALUE SUMMARY", DARK_BLUE);  r += 1

    for label, pv_tv_row in [
        ("EV  —  GGM Method  (Sum of PV FCFFs + PV of GGM TV)",          PV_GGM_ROW),
        ("EV  —  Exit Multiple Method  (Sum of PV FCFFs + PV of Exit TV)", PV_EXIT_ROW),
    ]:
        ws.row_dimensions[r].height = 20
        fill_cell(r, "A", label, bold=True, fg=WHITE, bg=RESULT_GREEN, align="left")
        fill_cell(r, "B", f"=B{SUM_PV_ROW}+B{pv_tv_row}",
                  bold=True, fg=WHITE, bg=RESULT_GREEN)
        if label.startswith("EV  —  GGM"):
            EV_GGM_ROW = r
        else:
            EV_EXIT_ROW = r
        r += 1

    # Mid-point
    ws.row_dimensions[r].height = 22
    fill_cell(r, "A", "EV Range Mid-Point  (Average of GGM and Exit Multiple)",
              bold=True, fg=WHITE, bg=DARK_BLUE, align="left")
    fill_cell(r, "B", f"=(B{EV_GGM_ROW}+B{EV_EXIT_ROW})/2",
              bold=True, fg=WHITE, bg=DARK_BLUE)
    r += 1

    r += 1  # blank

    # Note
    ws.row_dimensions[r].height = 14
    nc = ws.cell(row=r, column=1,
        value=("Note: Blue = manual input. Green = formula (do not edit). "
               "FCFF = NOPAT + D&A − CapEx − ΔWC. "
               "ΔWC = WC(curr period) − WC(prior period). "
               "All formulas link directly to source data."))
    nc.font      = Font(name="Calibri", italic=True, color="595959", size=9)
    nc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if periods:
        ws.merge_cells(f"A{r}:{periods[-1]['dcf_col']}{r}")



# ─── Section 8: UPLOAD ENDPOINT ────────────────────────────────────────────

@app.post("/upload")
async def upload_and_value(
    file: UploadFile = File(...),
    company_name: str = Form(...),
    sector: str = Form(...),
    valuation_date: str = Form(...)
):
    contents = await file.read()

    # Parse valuation date
    val_date_info = parse_valuation_date(valuation_date)

    # Load the client workbook
    wb = openpyxl.load_workbook(io.BytesIO(contents))

    # Classify columns across all sheets
    column_map = classify_columns(wb, val_date_info)
    print(f"Column map: {json.dumps(column_map, indent=2, default=str)}")

    # Add WACC tab
    add_wacc_tab(wb)

    # Find financial rows in source sheets
    financial_map = find_financial_rows(wb, column_map)
    print(f"Financial map: {json.dumps(financial_map, indent=2, default=str)}")

    # Add DCF Valuation tab
    add_dcf_tab(wb, val_date_info, column_map, financial_map, company_name)

    # Save modified workbook to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = company_name.replace(" ", "_") + "_Valuation_Model.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
